import pandas as pd
import json
from openpyxl import load_workbook
import json
from pymongo import MongoClient,UpdateMany

client_list=[]

list_tenants={}
details={}
configList=[]

def create_or_load_excel(filename, sheet_name, columns):
    print(f'{filename} {sheet_name} {columns}')
    try:
        # Try loading the workbook
        workbook = load_workbook(filename)
        # If the sheet exists, return it
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    except FileNotFoundError:
        pass  # File not found, proceed to create a new one

    # Create a new workbook and sheet
    workbook = pd.ExcelWriter(filename, engine='openpyxl')
    workbook.book.create_sheet(title=sheet_name, index=0)
    workbook.save()

    # Add columns to the sheet
    df = pd.DataFrame(columns=columns)
    df.to_excel(workbook, sheet_name=sheet_name, index=False)
    workbook.save()

    return workbook[sheet_name]

def insert_data(sheet, data):
    # Load existing data
    existing_data = pd.read_excel(sheet)
    # Append new data
    new_data = pd.DataFrame(data)
    updated_data = existing_data.append(new_data, ignore_index=True)
    # Write back to the sheet
    updated_data.to_excel(sheet, index=False)



def getTenants():
    with open('multi-tenancy.yml','r') as reader:
        for line in reader.readlines():
            dbip=''
            details={}
            if 'tenantName' in line:
                tenantName=line.strip('tenantName: ').split(':')[1]
                details['tenant']=tenantName
            if 'uri' in line:
                uri=line.strip('uri: ')
                dbip=line.split(':')[3].split('@')[1]
                details['dburi']=uri
                global list_tenants
                list_tenants.setdefault(dbip,[]).append(details)
                details={}





if __name__ == "__main__":
    # Specify the Excel file, sheet name, columns, and JSON data
    excel_file = "colleges_data.xlsx"
    column_names = ["Name", "Email", "Phone","Department", "Degree", "Year of Passing", "College ID"]


    getTenants()
    for  serverip in list_tenants:
        try:
            for  dburi in list_tenants[serverip]:
                try:
                    configDetails={}
                    dbname=dburi['dburi'].split('/')[3].split('?')[0].strip()
                    client=MongoClient(dburi['dburi'].strip())
                    db=client[dbname]
                    college_col=db['college']
                    cursor=college_col.find({})
                    for document in cursor:
                        collage_name = document['name']
                    # dbname_without_prefix = str(dbname)[len("dhi_"):]
                    # print(f"Client list size {len(client_list)} {dbname_without_prefix}")
                    # realm = next((client for client in client_list if client.get("resource") == dbname_without_prefix), None)
                    # print(f"Found Realm: {realm} ")
                    # realmcol.remove()
                    # realmcol.insert(realm)
                    print(collage_name)
                    usercol=db['dhi_user']
                    user_list=usercol.find({"userType":"ALUMNI"},{"name":1,"email":1,"mobile":1,"deptName":1,"degreeId":1,"academicYear":1,"tenantId":1})

                    # update_many = UpdateMany({}, {"$set":{"tenantId":dbname_without_prefix}})
                    # result = usercol.bulk_write([update_many])
                    # print("Modified:", result.modified_count)
                    sheet = create_or_load_excel(excel_file, collage_name, column_names)
                    for user in user_list:

    # Insert data into the sheet
                        insert_data(sheet, [user])

                except Exception as e:
                    print(e)
        except Exception as e:
            print(e)


    # # JSON data to be inserted
    # json_data = {
    #     "Name": "John Doe",
    #     "Email": "john.doe@example.com",
    #     "Phone": "123-456-7890",
    #     "Degree": "Computer Science",
    #     "Year of Passing": 2022,
    #     "College": "ABC University"
    # }

    # # Create or load the Excel sheet
  

    # print("Data inserted successfully.")
