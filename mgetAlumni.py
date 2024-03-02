import pandas as pd
from openpyxl import Workbook, load_workbook
from pymongo import MongoClient, UpdateMany

client_list = []
list_tenants = {}
details = {}
configList = {}


def create_or_load_excel(filename, sheet_name, columns):
    try:
        # Try loading the workbook
        workbook = load_workbook(filename)
        # If the sheet exists, return it
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    except FileNotFoundError:
        pass  # File not found, proceed to create a new one

    # Create a new workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # Add headers to the sheet
    sheet.append(columns)

    # Save the workbook
    workbook.save(filename)

    return sheet


def insert_data(sheet, data):
    # Append new data to the sheet
    sheet.append(data)
    # Save the workbook
    sheet.book.save(sheet.book.filename)


def getTenants():
    with open('multi-tenancy.yml', 'r') as reader:
        for line in reader.readlines():
            dbip = ''
            details = {}
            if 'tenantName' in line:
                tenantName = line.strip('tenantName: ').split(':')[1]
                details['tenant'] = tenantName
            if 'uri' in line:
                uri = line.strip('uri: ')
                dbip = line.split(':')[3].split('@')[1]
                details['dburi'] = uri
                global list_tenants
                list_tenants.setdefault(dbip, []).append(details)
                details = {}


if __name__ == "__main__":
    # Specify the Excel file, sheet name, columns, and JSON data
    excel_file = "colleges_data.xlsx"
    column_names = ["Name", "Email", "Phone", "Department", "Degree", "Year of Passing", "College ID"]

    getTenants()
    for serverip in list_tenants:
        try:
            for dburi in list_tenants[serverip]:
                try:
                    configDetails = {}
                    dbname = dburi['dburi'].split('/')[3].split('?')[0].strip()
                    client = MongoClient(dburi['dburi'].strip())
                    db = client[dbname]
                    college_col = db['college']
                    cursor = college_col.find({})
                    for document in cursor:
                        collage_name = document['name']
                    print(collage_name)
                    usercol = db['dhi_user']
                    user_list = usercol.find({"userType": "ALUMNI"},
                                             {"name": 1, "email": 1, "mobile": 1, "deptName": 1, "degreeId": 1,
                                              "academicYear": 1, "tenantId": 1})

                    sheet = create_or_load_excel(excel_file, collage_name, column_names)
                    for user in user_list:
                        insert_data(sheet, [user["name"], user["email"], user["mobile"],
                                            user["deptName"], user["degreeId"], user["academicYear"], user["tenantId"]])

                except Exception as e:
                    print(e)
        except Exception as e:
            print(e)

    print("Data inserted successfully.")

