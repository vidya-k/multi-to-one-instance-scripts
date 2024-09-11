from openpyxl import load_workbook, Workbook
from pymongo import MongoClient
import yaml
import bcrypt
import bson
# Load the multitenancy configuration from YAML
def load_multitenancy_config(yaml_file):
    with open(yaml_file, 'r') as file:
        return yaml.safe_load(file)

# Connect to MongoDB based on the tenantId from the YAML config
def connect_to_db(tenant_id, config):
    tenantDataList=config.get("tenantDataList")
    print(tenantDataList)
    filtered_tenants = [item for item in tenantDataList if item["tenantName"] == tenant_id]
    if filtered_tenants:
        uri=filtered_tenants[0]["uri"]
        mongoUri='{}/{}'.format(uri.split(',')[0],uri.split(',')[2].split('/')[1].split('?')[0])
        dbName='{}'.format(uri.split(',')[2].split('/')[1].split('?')[0])
        client = MongoClient(mongoUri)
        db = client[dbName]
        return db
    else:
        raise ValueError(f"Tenant ID {tenant_id} not found in the configuration")

# Create role objects from the roles and roleNames columns
def create_role_objects(roles_list, role_names_list):
    role_objects = []
    roleId= 5
    basicRole= True
    for role, role_name in zip(roles_list, role_names_list):
        role_objects.append({
            "roleDisplayName": role,
            "roleName": role_name,
            "roleId":roleId,
            "priority":roleId,
            "basicRole":basicRole
        })
        roleId=roleId+1
        basicRole=False
    return role_objects

# Generate a bcrypt password hash
def generate_bcrypt_password(password):
    salt = bcrypt.gensalt()
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed_password.decode('utf-8')

# Create a new user with bcrypt password and save to the database
def generate_password(user):
    # Generate a random password (you can enhance this logic to meet specific requirements)
    new_password = "random_password"  # Replace with password generation logic

    # Hash the password using bcrypt
    hashed_password = generate_bcrypt_password(new_password)
    user['password']=hashed_password
    # Return the new password for writing to Excel
    return new_password

# Process the Excel file and update or create users
def process_excel(file_path, yaml_config, output_excel_path):
    # Load the Excel workbook
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Load the multitenancy config
    config = load_multitenancy_config(yaml_config)

    # Create a new workbook to store the generated passwords
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.append(["College","New Email", "Generated Password"])

    # Iterate through rows (starting from row 2 to skip headers)
    for row in sheet.iter_rows(min_row=2, values_only=True):
      if any(row):
        tenant_id, roles, role_names, existing_email, new_email = row
        # Split roles and role_names by comma
        roles_list = roles.split(',')
        role_names_list = role_names.split(',')

        # Connect to the tenant's MongoDB
        db = connect_to_db(tenant_id, config)

        # Fetch the user by existing email
        user_collection = db["dhi_user"]  # Assuming 'users' collection
        user = user_collection.find_one({"email": existing_email})
        new_user = user_collection.find_one({"email": new_email})

        if user and new_user == None:
            # Update existing user
            role_objects = create_role_objects(roles_list, role_names_list)
            user["_id"]=bson.ObjectId()
            user['roles'] = role_objects
            user['email'] = new_email
            new_password = generate_password(user)
            user_collection.insert(user)
            print(f"Updated user: {existing_email} to {new_email} with new roles.")
         # Write new email and generated password to the output Excel
            output_sheet.append([tenant_id,new_email, new_password])
        else:
             print(f"user: {new_email} already present.")
           

    # Save the new Excel file with generated passwords
    output_workbook.save(output_excel_path)

# Example usage
if __name__ == "__main__":
    process_excel("user_creation.xlsx", "multi-tenancy.yml", "output_passwords.xlsx")
