import pandas as pd
from openpyxl import load_workbook
from pymongo import MongoClient, UpdateMany
import yaml

client_list = []


def create_or_load_excel(filename, sheet_name, columns):
    try:
        # Try loading the workbook
        workbook = load_workbook(filename)
        # If the sheet exists, return it
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    except FileNotFoundError:
        pass  # File not found, proceed to create a new one

    # Create a new workbook and sheet
    workbook = pd.ExcelWriter(filename, engine='openpyxl')
    workbook.book.create_sheet(title=sheet_name, index=0)
    workbook.save()

    # Add columns to the sheet
    df = pd.DataFrame(columns=columns)
    df.to_excel(workbook, sheet_name=sheet_name, index=False)
    workbook.save()

    return workbook[sheet_name]


def insert_data(sheet, data):
    # Load existing data
    existing_data = pd.read_excel(sheet)
    # Append new data
    new_data = pd.DataFrame(data)
    updated_data = existing_data.append(new_data, ignore_index=True)
    # Write back to the sheet
    updated_data.to_excel(sheet, index=False)


def get_tenants_from_yaml(file_path):
    tenants = {}
    data = read_yaml_file(file_path)

    if data and 'tenantDataList' in data:
        for tenant_data in data['tenantDataList']:
            tenant_name = tenant_data.get('tenantName')
            uri = tenant_data.get('uri')

            if tenant_name and uri:
                tenants[tenant_name] = uri

    return tenants


def read_yaml_file(file_path):
    with open(file_path, 'r') as file:
        try:
            data = yaml.safe_load(file)
            return data
        except yaml.YAMLError as e:
            print(f"Error reading YAML file: {e}")
            return None


if __name__ == "__main__":
    excel_file = "colleges_data.xlsx"
    column_names = ["Name", "Email", "Phone", "Department", "Degree", "Year of Passing", "College ID"]

    file_path = 'multi-tenancy.yml'
    list_tenants = get_tenants_from_yaml(file_path)

    if list_tenants:
        for serverip, details in list_tenants.items():
            print(f'{serverip} and {details}')
            try:
                for dburi in details:
                    try:
                        config_details = {}
                        dbname = dburi['dburi'].split('/')[3].split('?')[0].strip()
                        client = MongoClient(dburi['dburi'].strip())
                        db = client[dbname]
                        college_col = db['college']
                        cursor = college_col.find({})
                        for document in cursor:
                            collage_name = document['name']
                        print(collage_name)
                        user_col = db['dhi_user']
                        user_list = user_col.find({"userType": "ALUMNI"},
                                                  {"name": 1, "email": 1, "mobile": 1, "deptName": 1, "degreeId": 1,
                                                   "academicYear": 1, "tenantId": 1})

                        sheet = create_or_load_excel(excel_file, collage_name, column_names)
                        for user in user_list:
                            # Insert data into the sheet
                            insert_data(sheet, [user])

                    except Exception as e:
                        print(e)
            except Exception as e:
                print(e)

        print("Data inserted successfully.")
    else:
        print("Failed to read YAML file.")

